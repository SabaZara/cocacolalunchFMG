"""Report queries and file builders (CSV / XLSX). All file content is Georgian.

People are identified by card_id (names hidden for now). Two attendance shapes:
  * single day  → each active card marked "ჭამა" / "არ უჭამია" + a summary count.
  * multi-day   → days-attended out of days-in-range + a status.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlmodel import Session, func, select

from .config import get_settings
from .models import Person, Scan
from .timeutil import local_date_for, local_time_str, to_local, utc_now

# Georgian labels reused across reports/exports.
L_CARD_ID = "ბარათის ID"
L_STATUS = "სტატუსი"
L_ATE = "ჭამა"
L_NOT_ATE = "არ უჭამია"
L_TIME = "დრო"
L_DATE = "თარიღი"
L_COUNT = "რაოდენობა"
L_MEALS = "კვების რაოდენობა"
L_LIMIT = "დღიური ლიმიტი"
L_DAYS_ATTENDED = "დასწრების დღეები"
L_DAYS_IN_RANGE = "დღეები პერიოდში"
L_TOTAL_ATE = "სულ ნაჭამი"
L_TOTAL_ACTIVE = "აქტიური ბარათები"
L_PERIOD = "პერიოდი"
L_MEAL1 = "პირველი კვება"
L_MEAL2 = "მეორე კვება"
L_SUM = "ჯამი"
L_WINDOW = "კვება"


# --------------------- real-time correction + meal windows ------------------ #
def _real_local(dt_utc: datetime) -> datetime:
    """Recorded scan time -> REAL local time.

    The kiosk PC's clock runs ahead by KIOSK_CLOCK_AHEAD_MINUTES (default 14),
    so every stored timestamp is that much later than reality. All report
    times/buckets use this corrected value. (The kiosk screen itself is not
    adjusted — only reports.)
    """
    s = get_settings()
    return to_local(dt_utc, s.tz) - timedelta(minutes=s.kiosk_clock_ahead_min)


def _hhmm_to_min(hhmm: str) -> int:
    """'HH:MM' -> minutes since midnight; '24:00' -> 1440."""
    try:
        h, m = hhmm.split(":")
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return 0


def _split_min() -> int:
    """Admin-editable day-split boundary (minutes since midnight)."""
    from .app_config import get_meal_split
    return _hhmm_to_min(get_meal_split())


def _window_of(rl: datetime) -> int | None:
    """Which meal window a REAL local datetime falls in.

    Single day-split model: before the split = window 1 (პირველი კვება),
    at/after the split = window 2 (მეორე კვება). Every meal counts in one of
    the two, so this never returns None (kept Optional for API stability).
    """
    t = rl.hour * 60 + rl.minute
    return 1 if t < _split_min() else 2


# ------------------------------ summaries ---------------------------------- #
def today_summary(session: Session) -> dict:
    tz = get_settings().tz
    today = local_date_for(utc_now(), tz)
    # total meals claimed today (one row per meal)
    meals = session.exec(
        select(func.count()).select_from(Scan).where(Scan.local_date == today)
    ).one()
    # distinct PEOPLE who ate at least once today
    people_ate = session.exec(
        select(func.count(func.distinct(Scan.person_id))).where(Scan.local_date == today)
    ).one()
    active = session.exec(
        select(func.count()).select_from(Person).where(Person.active == True)  # noqa: E712
    ).one()
    return {
        "date": today.isoformat(),
        # `ate` kept = distinct people (back-compat); `meals` = total meals.
        "ate": int(people_ate),
        "people_ate": int(people_ate),
        "meals": int(meals),
        "active": int(active),
        # people who have NOT eaten at all yet today
        "remaining": max(int(active) - int(people_ate), 0),
    }


def daily_counts(session: Session, frm: date, to: date) -> list[dict]:
    rows = session.exec(
        select(Scan.local_date, func.count())
        .where(Scan.local_date >= frm, Scan.local_date <= to)
        .group_by(Scan.local_date)
        .order_by(Scan.local_date)
    ).all()
    by_date = {d: int(c) for d, c in rows}
    out: list[dict] = []
    cur = frm
    while cur <= to:
        out.append({"date": cur.isoformat(), "count": by_date.get(cur, 0)})
        cur += timedelta(days=1)
    return out


@dataclass
class DayRow:
    card_id: str
    count: int
    times: list[str]


def day_detail(session: Session, day: date, window: int | None = None) -> tuple[list[DayRow], int, int]:
    """Who ate on a given local day, GROUPED per card: meal count + REAL times.

    Times are clock-corrected (see _real_local). If `window` is 1 or 2, only
    meals inside that meal window are included. Returns (rows, w1_total,
    w2_total) where the totals are for the WHOLE day regardless of the filter.
    """
    scans = session.exec(
        select(Scan).where(Scan.local_date == day).order_by(Scan.scanned_at)
    ).all()
    grouped: dict[str, list[str]] = {}
    w1 = w2 = 0
    for s in scans:
        rl = _real_local(s.scanned_at)
        w = _window_of(rl)
        if w == 1:
            w1 += 1
        elif w == 2:
            w2 += 1
        if window in (1, 2) and w != window:
            continue
        grouped.setdefault(s.card_id, []).append(rl.strftime("%H:%M:%S"))
    rows = [DayRow(card_id=cid, count=len(times), times=times)
            for cid, times in grouped.items()]
    # chronological: earliest meal first ("HH:MM:SS" sorts lexicographically),
    # so inside the პირველი/მეორე filters the list reads in eating order.
    rows.sort(key=lambda r: (r.times[0] if r.times else "99:99:99", r.card_id))
    return rows, w1, w2


def detail_rows(session: Session, frm: date, to: date) -> list[dict]:
    """Flat detail rows for any range: date, card_id, REAL time, meal window."""
    scans = session.exec(
        select(Scan)
        .where(Scan.local_date >= frm, Scan.local_date <= to)
        .order_by(Scan.local_date, Scan.scanned_at)
    ).all()
    out = []
    for s in scans:
        rl = _real_local(s.scanned_at)
        w = _window_of(rl)
        out.append({
            "date": s.local_date.isoformat(),
            "card_id": s.card_id,
            "time": rl.strftime("%H:%M:%S"),
            "window": L_MEAL1 if w == 1 else (L_MEAL2 if w == 2 else "—"),
        })
    return out


# ------------------------ quantitative (marketing) ------------------------- #
def quantitative(session: Session, frm: date, to: date) -> dict:
    """Per-day meal counts split into the two meal windows + grand total.

    Buckets by REAL (clock-corrected) time of day. ჯამი = window1 + window2.
    """
    scans = session.exec(
        select(Scan).where(Scan.local_date >= frm, Scan.local_date <= to)
    ).all()
    per: dict[date, list[int]] = {}
    for s in scans:
        w = _window_of(_real_local(s.scanned_at))
        if w is None:
            continue
        bucket = per.setdefault(s.local_date, [0, 0])
        bucket[w - 1] += 1

    rows = []
    g1 = g2 = 0
    cur = frm
    while cur <= to:
        m1, m2 = per.get(cur, [0, 0])
        rows.append({"date": cur.isoformat(), "meal1": m1, "meal2": m2,
                     "total": m1 + m2})
        g1 += m1
        g2 += m2
        cur += timedelta(days=1)

    from .app_config import get_meal_split
    split = get_meal_split()
    return {
        "from": frm.isoformat(),
        "to": to.isoformat(),
        "rows": rows,
        "grand": {"meal1": g1, "meal2": g2, "total": g1 + g2},
        "labels": {"meal1": L_MEAL1, "meal2": L_MEAL2, "total": L_SUM},
        "split": split,
        "windows": {"meal1": f"00:00–{split}", "meal2": f"{split}–24:00"},
    }


def quantitative_csv(data: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([L_PERIOD, f'{data["from"]} — {data["to"]}'])
    w.writerow([L_DATE, L_MEAL1, L_MEAL2, L_SUM])
    for r in data["rows"]:
        w.writerow([r["date"], r["meal1"], r["meal2"], r["total"]])
    g = data["grand"]
    w.writerow([L_SUM, g["meal1"], g["meal2"], g["total"]])
    return buf.getvalue().encode("utf-8-sig")


def quantitative_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "რაოდენობრივი"
    ws.append([L_PERIOD, f'{data["from"]} — {data["to"]}'])
    ws.append([L_DATE, L_MEAL1, L_MEAL2, L_SUM])
    for col in range(1, 5):
        c = ws.cell(row=2, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for r in data["rows"]:
        ws.append([r["date"], r["meal1"], r["meal2"], r["total"]])
    g = data["grand"]
    ws.append([L_SUM, g["meal1"], g["meal2"], g["total"]])
    for col in range(1, 5):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
    _autofit(ws, [16, 16, 16, 12])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------- attendance matrix ----------------------------- #
def _days_in_range(frm: date, to: date) -> int:
    return (to - frm).days + 1


def attendance(session: Session, frm: date, to: date) -> dict:
    """Compute attendance per ACTIVE card over [frm, to]."""
    active_people = session.exec(
        select(Person).where(Person.active == True).order_by(Person.card_id)  # noqa: E712
    ).all()

    # person_id -> set of local_dates attended; and total meal count in range
    scans = session.exec(
        select(Scan.person_id, Scan.local_date).where(
            Scan.local_date >= frm, Scan.local_date <= to
        )
    ).all()
    attended: dict[int, set] = {}
    meal_count: dict[int, int] = {}
    for pid, d in scans:
        attended.setdefault(pid, set()).add(d)
        meal_count[pid] = meal_count.get(pid, 0) + 1

    days = _days_in_range(frm, to)
    single = days == 1
    rows = []
    total_ate = 0
    for p in active_people:
        n = len(attended.get(p.id, set()))
        if n > 0:
            total_ate += 1 if single else 0
        rows.append(
            {
                "card_id": p.card_id,
                "days_attended": n,
                "attended": n > 0,
                "meals": meal_count.get(p.id, 0),       # total meals in range
                "daily_limit": int(p.daily_limit),
            }
        )
    if not single:
        total_ate = sum(1 for r in rows if r["days_attended"] > 0)

    return {
        "from": frm.isoformat(),
        "to": to.isoformat(),
        "days": days,
        "single_day": single,
        "rows": rows,
        "total_active": len(active_people),
        "total_ate": total_ate,
    }


# ------------------------------ CSV builders ------------------------------- #
def detail_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([L_DATE, L_CARD_ID, L_TIME, L_WINDOW])
    for r in rows:
        w.writerow([r["date"], r["card_id"], r["time"], r.get("window", "")])
    # utf-8-sig so Excel opens Georgian correctly.
    return buf.getvalue().encode("utf-8-sig")


def attendance_csv(data: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    if data["single_day"]:
        w.writerow([L_CARD_ID, L_STATUS, L_MEALS, L_LIMIT])
        for r in data["rows"]:
            w.writerow([r["card_id"], L_ATE if r["attended"] else L_NOT_ATE,
                        r.get("meals", 0), r.get("daily_limit", "")])
        w.writerow([])
        w.writerow([L_TOTAL_ATE, data["total_ate"]])
        w.writerow([L_TOTAL_ACTIVE, data["total_active"]])
    else:
        w.writerow([L_CARD_ID, L_DAYS_ATTENDED, L_DAYS_IN_RANGE, L_STATUS])
        for r in data["rows"]:
            status = L_ATE if r["days_attended"] > 0 else L_NOT_ATE
            w.writerow([r["card_id"], r["days_attended"], data["days"], status])
        w.writerow([])
        w.writerow([L_TOTAL_ATE, data["total_ate"]])
        w.writerow([L_TOTAL_ACTIVE, data["total_active"]])
    return buf.getvalue().encode("utf-8-sig")


# ------------------------------ XLSX builders ------------------------------ #
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ATE_FILL = PatternFill("solid", fgColor="C6EFCE")
_NOT_FILL = PatternFill("solid", fgColor="FFC7CE")


def _style_header(ws, ncols: int) -> None:  # noqa: ANN001
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def _autofit(ws, widths: list[int]) -> None:  # noqa: ANN001
    from openpyxl.utils import get_column_letter

    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt


def _force_text(ws) -> None:  # noqa: ANN001
    """Force the card-id column (col 1) to text so leading zeros are preserved."""
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "@"


def detail_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "დეტალები"
    ws.append([L_DATE, L_CARD_ID, L_TIME, L_WINDOW])
    for r in rows:
        ws.append([r["date"], str(r["card_id"]), r["time"], r.get("window", "")])
    _style_header(ws, 4)
    _autofit(ws, [14, 22, 12, 16])
    # card_id is column 2 here; force it to text.
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "@"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def attendance_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "დასწრება"

    if data["single_day"]:
        ws.append([L_CARD_ID, L_STATUS, L_MEALS, L_LIMIT])
        _style_header(ws, 4)
        for r in data["rows"]:
            status = L_ATE if r["attended"] else L_NOT_ATE
            ws.append([str(r["card_id"]), status, r.get("meals", 0), r.get("daily_limit", "")])
            ws.cell(row=ws.max_row, column=2).fill = (
                _ATE_FILL if r["attended"] else _NOT_FILL
            )
        _autofit(ws, [22, 16, 18, 16])
    else:
        ws.append([L_CARD_ID, L_DAYS_ATTENDED, L_DAYS_IN_RANGE, L_STATUS])
        _style_header(ws, 4)
        for r in data["rows"]:
            attended = r["days_attended"] > 0
            status = L_ATE if attended else L_NOT_ATE
            ws.append([str(r["card_id"]), r["days_attended"], data["days"], status])
            ws.cell(row=ws.max_row, column=4).fill = (
                _ATE_FILL if attended else _NOT_FILL
            )
        _autofit(ws, [22, 18, 16, 16])

    _force_text(ws)

    # Summary block below a blank row.
    ws.append([])
    ws.append([L_PERIOD, f'{data["from"]} — {data["to"]}'])
    ws.append([L_TOTAL_ATE, data["total_ate"]])
    ws.append([L_TOTAL_ACTIVE, data["total_active"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
